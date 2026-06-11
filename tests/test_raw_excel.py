"""Excel Raw Extractor 테스트.

openpyxl 로 임시 .xlsx 를 만든 뒤 :func:`extract_excel_raw` 가 physical_raw 를
어떻게 만드는지 검증한다. 기획의 '충전환경온도' 표준규격 표를 작은 형태로 재현해
병합셀/헤더/데이터 셀이 raw json 에 그대로 담기는지 확인한다.

openpyxl 미설치 환경에서는 전체 모듈을 skip 한다.
"""

from __future__ import annotations

import pytest

openpyxl = pytest.importorskip("openpyxl")

from contentcompare.raw import extract_raw, raw_to_json
from contentcompare.raw.excel_raw import extract_excel_raw


@pytest.fixture()
def standard_xlsx(tmp_path):
    """'충전환경온도' 예시를 본뜬 표준규격 표(병합 멀티헤더 + 데이터 1행)."""
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "StandardList"

    # 2~3행 멀티헤더. '정량규격' 이 F2:H2 로 가로 병합, 'SPEC' 은 I2:I3 세로 병합.
    ws["D2"] = "항목"
    ws.merge_cells("D2:E2")
    ws["F2"] = "정량규격"
    ws.merge_cells("F2:H2")
    ws["I2"] = "SPEC(중심치+-상하한치)"
    ws.merge_cells("I2:I3")
    ws["J2"] = "정성규격"
    ws.merge_cells("J2:J3")
    ws["K2"] = "단위"
    ws.merge_cells("K2:K3")

    ws["D3"] = "중분류"
    ws["E3"] = "소분류"
    ws["F3"] = "하한치"
    ws["G3"] = "중심치"
    ws["H3"] = "상한치"

    # 헤더 행은 굵게 + 노란 채움(스타일이 raw 에 담기는지 확인용).
    bold = Font(bold=True)
    yellow = PatternFill(fill_type="solid", fgColor="FFFFFF00")
    for addr in ("D2", "F2", "I2", "J2", "K2", "D3", "E3", "F3", "G3", "H3"):
        ws[addr].font = bold
        ws[addr].fill = yellow

    # 데이터 1행(17행 대신 4행으로 축약).
    ws["B4"] = 10
    ws["C4"] = "기본사양"
    ws["D4"] = "기본사양"
    ws["E4"] = "충전환경온도"
    ws["F4"] = -5
    ws["G4"] = 25
    ws["H4"] = 55
    ws["I4"] = "25+-30"
    ws["J4"] = "-5 to 5도씨 0.1C (4.55V)"
    ws["K4"] = "도씨"
    ws["G4"].number_format = "0.0"
    ws["E4"].comment = Comment("측정 조건 확인 필요", "tester")

    path = tmp_path / "standard.xlsx"
    wb.save(path)
    return str(path)


def _cell(sheet_dict, address):
    for c in sheet_dict["cells"]:
        if c["address"] == address:
            return c
    return None


def test_doc_root_shape(standard_xlsx):
    doc = extract_excel_raw(standard_xlsx).to_dict()
    assert doc["doc_type"] == "excel"
    assert doc["file_name"] == "standard.xlsx"
    assert len(doc["sheets"]) == 1
    assert doc["sheets"][0]["sheet_name"] == "StandardList"


def test_merged_regions_captured(standard_xlsx):
    sheet = extract_excel_raw(standard_xlsx).to_dict()["sheets"][0]
    ranges = {m["range"]: m["value"] for m in sheet["merged_cells"]}
    assert ranges["D2:E2"] == "항목"
    assert ranges["F2:H2"] == "정량규격"
    assert ranges["I2:I3"] == "SPEC(중심치+-상하한치)"


def test_merged_anchor_and_range_on_cells(standard_xlsx):
    sheet = extract_excel_raw(standard_xlsx).to_dict()["sheets"][0]
    f2 = _cell(sheet, "F2")
    assert f2["value"] == "정량규격"
    assert f2["merged_range"] == "F2:H2"
    assert f2["is_merged_anchor"] is True
    # 병합 영역 안의 빈 셀(G2/H2)은 값이 없으므로 cells 에 포함되지 않는다.
    assert _cell(sheet, "G2") is None


def test_data_cells_keep_value_and_type(standard_xlsx):
    sheet = extract_excel_raw(standard_xlsx).to_dict()["sheets"][0]
    e4 = _cell(sheet, "E4")
    assert e4["value"] == "충전환경온도"
    assert e4["row"] == 4 and e4["column"] == "E" and e4["col_index"] == 5

    f4 = _cell(sheet, "F4")
    assert f4["value"] == -5
    assert f4["value_type"] == "int"

    # 빈 셀(A 열 등)은 셀 목록에 없다.
    assert _cell(sheet, "A1") is None


def test_style_metadata_captured(standard_xlsx):
    sheet = extract_excel_raw(standard_xlsx).to_dict()["sheets"][0]
    f2 = _cell(sheet, "F2")
    assert f2["font_bold"] is True
    assert f2["fill_color"] == "FFFFFF00"

    g4 = _cell(sheet, "G4")
    assert g4["number_format"] == "0.0"

    e4 = _cell(sheet, "E4")
    assert e4["comment"] == "측정 조건 확인 필요"


def test_no_interpretation_only_physical(standard_xlsx):
    """raw 단계는 의미 해석을 하지 않는다 — entity/lower_limit 같은 키가 없어야 한다."""
    text = raw_to_json(extract_excel_raw(standard_xlsx))
    assert "entity" not in text
    assert "lower_limit" not in text
    # 대신 물리 정보 키는 존재.
    assert "merged_range" in text
    assert "address" in text


def test_dispatcher_routes_xlsx(standard_xlsx):
    doc = extract_raw(standard_xlsx).to_dict()
    assert doc["doc_type"] == "excel"


def test_json_serializable_and_korean_preserved(standard_xlsx):
    import json

    text = raw_to_json(extract_excel_raw(standard_xlsx))
    parsed = json.loads(text)  # 유효한 json 이어야 한다
    assert "충전환경온도" in text  # ensure_ascii=False 로 한글 보존
    assert parsed["sheets"][0]["sheet_name"] == "StandardList"
