"""Excel Raw Extractor — openpyxl 로 .xlsx 를 physical_raw 로 변환.

해석은 하지 않는다. 보이는 것만 담는다:
값 / 위치 / 병합영역 / 숫자서식 / 굵게·글자크기 / 채움색 / 코멘트 / 숨김.

왜 openpyxl 인가
----------------
기존 ``readers/excel_reader.py`` 는 사내 제약(COM)으로 xlwings 를 쓰지만, raw
추출 단계는 **Office 없이도** 돌아가야 개발·테스트·CI 가 쉽다. openpyxl 은
크로스플랫폼이고 병합/서식/코멘트 메타데이터를 그대로 읽을 수 있다. 추후 COM
백엔드가 필요하면 같은 :class:`RawExcelDocument` 를 채우는 다른 함수만 추가하면
된다(모델은 백엔드 무관).

openpyxl 은 선택적 의존성이므로 import 를 함수 안으로 지연한다.
"""

from __future__ import annotations

import logging
import os
from datetime import date, datetime, time
from typing import Any, Optional

from .models import RawCell, RawExcelDocument, RawMergedRegion, RawSheet

logger = logging.getLogger(__name__)


def _value_type(value: Any) -> str:
    """raw json 에 남길 값 타입 이름. datetime 계열은 별도 라벨."""
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, datetime):
        return "datetime"
    if isinstance(value, date):
        return "date"
    if isinstance(value, time):
        return "time"
    return type(value).__name__


def _json_safe(value: Any) -> Any:
    """json 직렬화 가능한 형태로 변환(날짜는 ISO 문자열). 숫자/문자는 그대로."""
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return value


def _argb(color: Any) -> Optional[str]:
    """openpyxl Color → ARGB 문자열. 자동/테마색 등은 None."""
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    # rgb 가 'FFFFFF00' 같은 8자리 hex 문자열일 때만 의미가 있다.
    if isinstance(rgb, str) and len(rgb) in (6, 8) and rgb not in ("00000000",):
        return rgb
    return None


def extract_excel_raw(path: str) -> RawExcelDocument:
    """xlsx 파일 경로 → :class:`RawExcelDocument`.

    각 시트의 비어있지 않은 셀만 :class:`RawCell` 로 만든다(빈 셀 제외).
    병합 영역은 좌상단 앵커 셀에 ``is_merged_anchor=True`` 로 표시하고, 영역 내
    모든 셀에는 ``merged_range`` 를 기록한다(앵커가 아닌 셀의 값은 openpyxl 에서
    보통 None 이므로 자동으로 cells 에 빠진다 — 정보 손실 없음).
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - 환경 의존
        raise RuntimeError(
            "openpyxl 이 필요합니다. pip install openpyxl"
        ) from exc

    file_name = os.path.basename(path)
    logger.info("[RawExcel] 열기: %s", os.path.abspath(path))
    # data_only=False: 수식 문자열을 보존(원문). 캐시된 계산값이 필요하면 별도 옵션.
    wb = load_workbook(path, data_only=False)

    doc = RawExcelDocument(file_name=file_name)
    for ws in wb.worksheets:
        doc.sheets.append(_extract_sheet(ws, get_column_letter))
    logger.info("[RawExcel] 완료: 시트 %d개", len(doc.sheets))
    return doc


def _extract_sheet(ws, get_column_letter) -> RawSheet:
    """openpyxl Worksheet → RawSheet."""
    # 셀 주소 → 병합영역 문자열 매핑(빠른 조회용).
    merged_lookup: dict[str, str] = {}
    merged_regions: list[RawMergedRegion] = []
    for mr in sorted(ws.merged_cells.ranges, key=lambda r: (r.min_row, r.min_col)):
        rng = str(mr)
        anchor = ws.cell(row=mr.min_row, column=mr.min_col)
        merged_regions.append(
            RawMergedRegion(range=rng, value=_json_safe(anchor.value))
        )
        for row in range(mr.min_row, mr.max_row + 1):
            for col in range(mr.min_col, mr.max_col + 1):
                merged_lookup[f"{get_column_letter(col)}{row}"] = rng

    hidden_cols = {
        letter
        for letter, dim in ws.column_dimensions.items()
        if getattr(dim, "hidden", False)
    }
    hidden_rows = {
        idx for idx, dim in ws.row_dimensions.items() if getattr(dim, "hidden", False)
    }

    cells: list[RawCell] = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            addr = cell.coordinate
            merged_range = merged_lookup.get(addr)
            is_anchor = bool(
                merged_range and merged_range.split(":")[0] == addr
            )
            fmt = cell.number_format
            font = cell.font
            cells.append(
                RawCell(
                    address=addr,
                    row=cell.row,
                    column=get_column_letter(cell.column),
                    col_index=cell.column,
                    value=_json_safe(cell.value),
                    value_type=_value_type(cell.value),
                    number_format=fmt if fmt and fmt != "General" else None,
                    merged_range=merged_range,
                    is_merged_anchor=is_anchor,
                    font_bold=bool(font.bold) if font is not None else None,
                    font_size=float(font.size)
                    if font is not None and font.size is not None
                    else None,
                    fill_color=_fill_color(cell),
                    comment=cell.comment.text if cell.comment is not None else None,
                )
            )

    dims = ws.dimensions  # 예: 'B2:P17' 또는 'A1'
    min_row = ws.min_row or 1
    max_row = ws.max_row or 1
    min_col = ws.min_column or 1
    max_col = ws.max_column or 1

    return RawSheet(
        sheet_name=ws.title,
        used_range=dims if cells else None,
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        merged_cells=merged_regions,
        cells=cells,
        hidden=(ws.sheet_state != "visible"),
    )


def _fill_color(cell) -> Optional[str]:
    """셀 배경 채움색(solid 패턴만). 채움 없음/패턴없음이면 None."""
    fill = getattr(cell, "fill", None)
    if fill is None or getattr(fill, "patternType", None) != "solid":
        return None
    return _argb(getattr(fill, "fgColor", None))
